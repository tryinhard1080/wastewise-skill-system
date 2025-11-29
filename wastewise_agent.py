"""
WasteWise Complete Suite agent – end-to-end scaffold
- Reads structured inputs (JSON), runs decision logic, and generates Excel (8 tabs) + HTML (6 tabs) outputs.
- Styling for HTML matches DEMO_Dashboard (green gradient, Inter, Chart.js).
"""

import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# --------------------
# Data models
# --------------------
@dataclass
class PropertyProfile:
    name: str
    units: int
    property_type: str  # garden, mid-rise, high-rise, mixed-use
    occupancy_pct: float
    status: str  # lease-up, stabilized, value-add/renovation
    has_compactor: bool
    has_valet: bool
    location: Optional[str] = None  # "City, State"


@dataclass
class EquipmentProfile:
    compactor_tons: Optional[float] = None
    compactor_max_days_between_pickups: Optional[int] = None
    dumpster_qty: Optional[int] = None
    dumpster_size: Optional[float] = None
    dumpster_freq_per_week: Optional[float] = None
    dual_compactors: bool = False


@dataclass
class Financials:
    monthly_cost: float
    pickup_cost_per_haul: float
    contamination_charges: float
    bulk_charges: float
    avg_monthly_overage: float
    avg_tons_per_haul: float
    overages_present: bool


@dataclass
class InvoiceEntry:
    month: str  # MM/YYYY
    invoice_number: str
    hauler: str
    pickup_fees: float
    disposal: float
    rental: float
    contamination: float
    bulk: float
    overages: float
    other: float
    units: int
    equipment_type: str


@dataclass
class ContractInfo:
    text: Optional[str] = None
    expiration_date: Optional[str] = None  # MM/DD/YYYY
    auto_renewal: Optional[str] = None
    termination_notice_days: Optional[int] = None
    rate_increase_terms: Optional[str] = None
    force_majeure_text: Optional[str] = None
    indemnification_text: Optional[str] = None
    service_level_text: Optional[str] = None
    liability_text: Optional[str] = None
    renewal_terms: Optional[str] = None
    termination_clause: Optional[str] = None
    provided: bool = False


@dataclass
class Recommendation:
    title: str
    detail: str
    monthly_savings: float = 0.0
    annual_savings: float = 0.0
    payback_months: Optional[float] = None
    priority: Optional[int] = None


# --------------------
# Helpers
# --------------------
def fmt_currency(value: float) -> str:
    return f"${value:,.2f}"


def fmt_percent(value: float) -> str:
    return f"{value:.1f}%"


def months_in_data(invoices: List[InvoiceEntry]) -> List[str]:
    uniq = sorted({inv.month for inv in invoices})
    return uniq


# --------------------
# Core agent
# --------------------
class WasteWiseAgent:
    def __init__(self):
        self.brand = "WasteWise by THE Trash Hub"
        self.contacts = {
            "primary": "Richard Bates (The Trash Hub)",
            "compactor_monitors": "Keith Conrad (DSQ Technologies) <keith.conrad@dsqtech.com>",
            "bulk_trash": "Cole Myers (Ally Waste) <cole@allywaste.com>",
        }

    # --- Core calculations ---
    def yards_per_door_compactor(self, total_tons: float, units: int) -> float:
        return (total_tons * 14.49) / units if units else 0

    def yards_per_door_compactor_alt(self, total_tons: float, units: int) -> float:
        return ((total_tons * 2000) / 138) / units if units else 0

    def yards_per_door_dumpster(self, qty: int, size: float, freq: float, units: int) -> float:
        return (qty * size * freq * 4.33) / units if units else 0

    def cost_per_door(self, monthly_cost: float, units: int) -> float:
        return monthly_cost / units if units else 0

    def annual_savings(self, monthly_savings: float) -> float:
        return monthly_savings * 12

    # --- Decision logic ---
    def compactor_optimization(
        self,
        equipment: EquipmentProfile,
        financials: Financials,
        monthly_pickup_savings: float,
        install_cost: float,
        monitoring_cost: float,
    ) -> Optional[Recommendation]:
        if financials.avg_tons_per_haul >= 7:
            return None
        if equipment.compactor_max_days_between_pickups and equipment.compactor_max_days_between_pickups > 14:
            return None
        if monthly_pickup_savings < 300:
            return None

        net_monthly = monthly_pickup_savings - (install_cost + monitoring_cost)
        if net_monthly <= 0:
            return None

        return Recommendation(
            title="Add Compactor Monitors",
            detail="Avg tons/haul <7 and interval ≤14 days; monitors target 8–9 tons/haul. Savings based on reduced pickups (pickup fees only).",
            monthly_savings=net_monthly,
            annual_savings=self.annual_savings(net_monthly),
        )

    def overage_strategy(self, financials: Financials, overage_frequency: str, extra_service_cost: float) -> str:
        annual_overage = financials.avg_monthly_overage * 12
        annual_added_service = extra_service_cost * 12

        if overage_frequency == "consistent":
            if annual_added_service < annual_overage:
                return "Add permanent service day; cheaper than overages."
            return "Overages cheaper than added service; keep status quo."

        if overage_frequency == "seasonal":
            return "Add seasonal service only during peak months."

        return "Investigate operations (valet distribution, compliance); consider larger equipment if needed."

    def contamination_plan(self, financials: Financials, total_spend: float) -> Optional[Recommendation]:
        rate = (financials.contamination_charges / total_spend) * 100 if total_spend else 0
        if rate <= 3:
            return None

        if rate > 5 and financials.contamination_charges > 150:
            detail = (
                "Full contamination reduction: signage $500, education $300, monitoring $50/month; "
                "expected 50% charge reduction."
            )
            monthly_savings = financials.contamination_charges * 0.5
        else:
            detail = "Light intervention: signage refresh and resident reminders."
            monthly_savings = financials.contamination_charges * 0.25

        return Recommendation(
            title="Contamination Reduction",
            detail=detail,
            monthly_savings=monthly_savings,
            annual_savings=self.annual_savings(monthly_savings),
        )

    def bulk_strategy(self, avg_monthly_bulk: float) -> Recommendation:
        if avg_monthly_bulk > 500:
            savings = (avg_monthly_bulk - 400) * 12
            return Recommendation(
                title="Switch to Bulk Subscription",
                detail="Average bulk > $500/month; subscription at $400/month lowers cost.",
                monthly_savings=avg_monthly_bulk - 400,
                annual_savings=savings,
            )
        if 300 <= avg_monthly_bulk <= 500:
            return Recommendation(
                title="Monitor Bulk Spend",
                detail="Borderline spend; monitor 3 months and prepare for subscription if trend increases.",
            )
        return Recommendation(
            title="Keep On-Demand Bulk",
            detail="On-demand pricing remains cost-effective.",
        )

    def service_level(self, financials: Financials, contamination_or_overages: bool) -> str:
        if contamination_or_overages and financials.avg_tons_per_haul < 6:
            return "Address contamination/overages before any reduction."

        if financials.avg_tons_per_haul >= 8 and financials.overages_present:
            return "Add service day (compactor near capacity)."
        if financials.avg_tons_per_haul < 6 and not financials.overages_present:
            return "Reduce pickup frequency (underutilized)."
        if 6 <= financials.avg_tons_per_haul < 8:
            return "Maintain current service."
        return "Maintain current service."

    def prioritize(self, recs: List[Recommendation]) -> List[Recommendation]:
        ranked = sorted(
            recs,
            key=lambda r: (-(r.annual_savings or 0), float("inf") if r.payback_months is None else r.payback_months),
        )
        for idx, rec in enumerate(ranked, start=1):
            rec.priority = idx
        return ranked

    # --- Validation ---
    def validate(self, profile: PropertyProfile, equipment: EquipmentProfile, invoices: List[InvoiceEntry]) -> List[str]:
        errors: List[str] = []
        if not profile.name:
            errors.append("Property name is required.")
        if profile.units <= 0:
            errors.append("Units must be positive.")
        if profile.status == "lease-up" and profile.occupancy_pct >= 90:
            errors.append("Lease-up status inconsistent with occupancy >= 90%.")
        if equipment.compactor_max_days_between_pickups and equipment.compactor_max_days_between_pickups > 14:
            errors.append("Compactor pickup interval exceeds 14-day threshold for optimization.")
        if not invoices:
            errors.append("At least one invoice is required.")
        return errors

    # --- Analysis pipeline ---
    def analyze(self, payload: Dict) -> Dict:
        profile = payload["property_profile"]
        invoices = [InvoiceEntry(**inv) for inv in payload.get("invoices", [])]
        haul_log = payload.get("haul_log", [])
        contract_meta = payload.get("contract_meta", {})

        months = months_in_data(invoices)
        months_count = max(1, len(months))
        units = profile["units"]

        total_spend = sum(
            inv.pickup_fees + inv.disposal + inv.rental + inv.contamination + inv.bulk + inv.overages + inv.other
            for inv in invoices
        )
        avg_monthly_spend = total_spend / months_count
        avg_cost_per_door = self.cost_per_door(avg_monthly_spend, units)

        total_contam = sum(inv.contamination for inv in invoices)
        total_bulk = sum(inv.bulk for inv in invoices)
        avg_monthly_bulk = total_bulk / months_count if months_count else 0
        avg_monthly_contam = total_contam / months_count if months_count else 0

        avg_tons_per_haul = profile.get("avg_tons_per_haul") or (
            sum(h.get("tons", 0) for h in haul_log) / len(haul_log) if haul_log else 0
        )
        pickup_cost_per_haul = (
            sum(inv.pickup_fees for inv in invoices) / len(invoices) if invoices else 0
        )

        hauls_per_month = (len(haul_log) / months_count) if haul_log else max(1, len(invoices) / months_count)
        target_tons = 8.0
        current_tons_per_month = avg_tons_per_haul * hauls_per_month
        projected_hauls = math.ceil(current_tons_per_month / target_tons) if target_tons else hauls_per_month
        hauls_saved = max(0, hauls_per_month - projected_hauls)
        monthly_pickup_savings = hauls_saved * pickup_cost_per_haul

        financials = Financials(
            monthly_cost=avg_monthly_spend,
            pickup_cost_per_haul=pickup_cost_per_haul,
            contamination_charges=avg_monthly_contam,
            bulk_charges=avg_monthly_bulk,
            avg_monthly_overage=payload.get("property_profile", {}).get("avg_monthly_overage", 0) or 0,
            avg_tons_per_haul=avg_tons_per_haul,
            overages_present=bool(payload.get("property_profile", {}).get("overages_present", False)),
        )
        equipment = EquipmentProfile(
            compactor_tons=profile.get("avg_tons_per_haul"),
            compactor_max_days_between_pickups=profile.get("max_days_between_pickups"),
            dumpster_qty=profile.get("dumpster_qty"),
            dumpster_size=profile.get("dumpster_size"),
            dumpster_freq_per_week=profile.get("dumpster_freq_per_week"),
            dual_compactors=bool(profile.get("dual_compactors", False)),
        )

        recs: List[Recommendation] = []
        compactor_rec = self.compactor_optimization(
            equipment=equipment,
            financials=financials,
            monthly_pickup_savings=monthly_pickup_savings,
            install_cost=200.0,
            monitoring_cost=50.0,
        )
        if compactor_rec:
            recs.append(compactor_rec)
        contamination_rec = self.contamination_plan(financials, total_spend=avg_monthly_spend)
        if contamination_rec:
            recs.append(contamination_rec)
        recs.append(self.bulk_strategy(avg_monthly_bulk))
        prioritized = self.prioritize(recs)

        total_annual_savings = sum(r.annual_savings for r in prioritized if r.annual_savings)
        service_guidance = self.service_level(
            financials, contamination_or_overages=financials.overages_present or financials.contamination_charges > 0
        )

        contract = ContractInfo(
            text=contract_meta.get("text"),
            expiration_date=contract_meta.get("expiration_date"),
            auto_renewal=contract_meta.get("auto_renewal"),
            termination_notice_days=contract_meta.get("termination_notice_days"),
            rate_increase_terms=contract_meta.get("rate_increase_terms"),
            force_majeure_text=contract_meta.get("force_majeure_text"),
            indemnification_text=contract_meta.get("indemnification_text"),
            service_level_text=contract_meta.get("service_level_text"),
            liability_text=contract_meta.get("liability_text"),
            renewal_terms=contract_meta.get("renewal_terms"),
            termination_clause=contract_meta.get("termination_clause"),
            provided=contract_meta.get("contract_provided", "").upper() == "YES",
        )

        return {
            "profile": profile,
            "equipment": equipment,
            "financials": financials,
            "invoices": invoices,
            "haul_log": haul_log,
            "contract": contract,
            "months": months,
            "metrics": {
                "total_spend": total_spend,
                "avg_monthly_spend": avg_monthly_spend,
                "avg_cost_per_door": avg_cost_per_door,
                "avg_tons_per_haul": avg_tons_per_haul,
                "pickup_cost_per_haul": pickup_cost_per_haul,
                "monthly_pickup_savings": monthly_pickup_savings,
                "bulk_avg": avg_monthly_bulk,
                "contam_avg": avg_monthly_contam,
                "contam_rate_pct": (avg_monthly_contam / avg_monthly_spend * 100) if avg_monthly_spend else 0,
                "yards_per_door": self.yards_per_door_compactor(sum(h.get("tons", 0) for h in haul_log), units)
                if profile.get("has_compactor")
                else self.yards_per_door_dumpster(
                    profile.get("dumpster_qty", 0) or 0,
                    profile.get("dumpster_size", 0) or 0,
                    profile.get("dumpster_freq_per_week", 0) or 0,
                    units,
                ),
                "hauls_per_month": hauls_per_month,
                "hauls_saved": hauls_saved,
            },
            "recommendations": prioritized,
            "service_guidance": service_guidance,
            "total_annual_savings": total_annual_savings,
        }

    # --- Excel builder ---
    def build_excel_workbook(self, analysis: Dict, path: Path) -> Path:
        wb = Workbook()
        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        bold = Font(bold=True)

        def make_sheet(title: str):
            ws = wb.create_sheet(title)
            return ws

        wb.remove(wb.active)

        ws = make_sheet("SUMMARY")
        ws.append(["WasteWise by THE Trash Hub"])
        ws.append(["Property", analysis["profile"]["name"]])
        ws.append(["Units", analysis["profile"]["units"]])
        ws.append(["Property Type", analysis["profile"]["property_type"]])
        ws.append(["Status", analysis["profile"]["status"]])
        ws.append(["Location", analysis["profile"].get("location", "")])
        ws.append(["Avg Cost/Door (monthly)", fmt_currency(analysis["metrics"]["avg_cost_per_door"])])
        ws.append(["Total Spend (period)", fmt_currency(analysis["metrics"]["total_spend"])])
        ws.append(["Annual Savings Opportunity", fmt_currency(analysis["total_annual_savings"])])
        for cell in ws["A1:A9"]:
            cell[0].font = bold

        ws = make_sheet("SUMMARY_FULL")
        ws.append([f"Potential to Reduce 2026 Trash Expense by {fmt_currency(analysis['total_annual_savings'])}"])
        ws.append([f"Service guidance: {analysis['service_guidance']}"])
        ws.append([f"Primary contact: {self.contacts['primary']}"])
        ws.append([f"Compactor monitors: {self.contacts['compactor_monitors']}"])
        ws.append([f"Bulk trash: {self.contacts['bulk_trash']}"])
        for cell in ws["A1:A5"]:
            cell[0].font = bold

        ws = make_sheet("EXPENSE_ANALYSIS")
        headers = [
            "Month",
            "Invoice #",
            "Hauler",
            "Pickup",
            "Disposal",
            "Rental",
            "Contam",
            "Bulk",
            "Overages",
            "Other",
            "Total",
            "Cost/Door",
            "Notes",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
        units = analysis["profile"]["units"]
        for inv in analysis["invoices"]:
            total = inv.pickup_fees + inv.disposal + inv.rental + inv.contamination + inv.bulk + inv.overages + inv.other
            ws.append(
                [
                    inv.month,
                    inv.invoice_number,
                    inv.hauler,
                    inv.pickup_fees,
                    inv.disposal,
                    inv.rental,
                    inv.contamination,
                    inv.bulk,
                    inv.overages,
                    inv.other,
                    total,
                    total / units if units else 0,
                    "",
                ]
            )

        ws = make_sheet("HAUL_LOG")
        ws.append(["pickup_date", "tons", "avg_tons_per_haul", "max_days_between_pickups", "yards_per_door", "notes"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
        for h in analysis["haul_log"]:
            ws.append(
                [
                    h.get("pickup_date"),
                    h.get("tons"),
                    h.get("avg_tons_per_haul"),
                    h.get("max_days_between_pickups"),
                    h.get("yards_per_door"),
                    h.get("notes", ""),
                ]
            )

        ws = make_sheet("OPTIMIZATION")
        ws.append(["Priority", "Recommendation", "Monthly Savings", "Annual Savings", "Detail"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
        for rec in analysis["recommendations"]:
            ws.append([rec.priority, rec.title, rec.monthly_savings, rec.annual_savings, rec.detail])

        ws = make_sheet("CONTRACT_TERMS")
        ws.append(["Clause", "Details", "Risk"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
        contract: ContractInfo = analysis["contract"]
        if contract.provided:
            ws.append(["Term & Renewal", f"Expires {contract.expiration_date}; {contract.auto_renewal}", "MED"])
            ws.append(["Rate Increases", contract.rate_increase_terms, "LOW"])
            ws.append(["Termination", f"{contract.termination_notice_days}-day notice; {contract.termination_clause}", "MED"])
            ws.append(["Liability", contract.liability_text, "MED"])
            ws.append(["Service Level", contract.service_level_text, "LOW"])
            ws.append(["Force Majeure", contract.force_majeure_text, "LOW"])
            ws.append(["Indemnification", contract.indemnification_text, "MED"])
        else:
            ws.append(["Contract", "No contract provided", "HIGH"])

        ws = make_sheet("REGULATORY_COMPLIANCE")
        ws.append(["Section", "Details"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
        ws.append(
            [
                "Confidence",
                "LOW — placeholder. Run targeted searches: '[City] [State] waste recycling ordinance', 'universal recycling multifamily', 'composting mandate commercial'.",
            ]
        )
        ws.append(
            [
                "Summary",
                f"Location: {analysis['profile'].get('location', 'N/A')}. Update with official .gov sources and list 3–5 licensed haulers.",
            ]
        )

        ws = make_sheet("INSTRUCTIONS")
        ws.append(["How to use"])
        ws.append(["1) Verify all invoice rows and haul log entries are included."])
        ws.append(["2) Confirm contract clauses are verbatim if provided."])
        ws.append(["3) Ensure regulatory tab is updated with official sources and confidence score."])

        wb.save(path)
        return path

    # --- HTML builder (DEMO style) ---
    def build_html_dashboard(self, analysis: Dict, path: Path) -> Path:
        m = analysis["metrics"]
        profile = analysis["profile"]
        recs = analysis["recommendations"]
        invoices = analysis["invoices"]
        haul = analysis["haul_log"]

        expense_labels = [inv.month for inv in invoices]
        expense_totals = [
            inv.pickup_fees + inv.disposal + inv.rental + inv.contamination + inv.bulk + inv.overages + inv.other
            for inv in invoices
        ]
        tonnage_labels = [h.get("pickup_date", "")[:5] for h in haul] if haul else []
        tonnage_totals = [h.get("tons", 0) for h in haul] if haul else []

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>WasteWise Dashboard - {profile['name']}</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.js"></script>
  <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.2/css/all.min.css" rel="stylesheet">
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: Inter, system-ui, Arial, sans-serif; background: #f7f7f7; color: #1B1B1B; }}
    .dashboard-header {{ padding: 24px; background: linear-gradient(135deg, #2C5F2D 0%, #1f4220 100%); color: white; box-shadow: 0 4px 12px rgba(0,0,0,0.15); }}
    .property-title {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
    .analysis-date {{ opacity: 0.9; font-size: 14px; margin-bottom: 20px; }}
    .kpi-row {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 20px; }}
    .kpi-card {{ background: rgba(255,255,255,0.1); border-radius: 12px; padding: 16px; text-align: center; backdrop-filter: blur(10px); transition: all 0.3s ease; }}
    .kpi-card:hover {{ background: rgba(255,255,255,0.15); transform: translateY(-2px); }}
    .kpi-card.primary {{ background: #4A7C4E; border: 2px solid #90EE90; box-shadow: 0 4px 12px rgba(144, 238, 144, 0.3); }}
    .kpi-label {{ font-size: 11px; opacity: 0.85; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px; }}
    .kpi-value {{ font-size: 24px; font-weight: 700; }}
    .kpi-card.primary .kpi-value {{ color: #90EE90; }}
    .tab-navigation {{ display: flex; gap: 8px; flex-wrap: wrap; }}
    .tab-btn {{ background: rgba(255,255,255,0.15); color: white; border: 0; padding: 12px 20px; border-radius: 8px; cursor: pointer; font-size: 14px; font-weight: 500; transition: all 0.2s; backdrop-filter: blur(10px); }}
    .tab-btn:hover {{ background: rgba(255,255,255,0.25); transform: translateY(-1px); }}
    .tab-btn.active {{ background: #90EE90; color: #1B1B1B; font-weight: 600; box-shadow: 0 4px 12px rgba(144, 238, 144, 0.3); }}
    .container {{ padding: 24px; max-width: 1400px; margin: 0 auto; }}
    .tab-content {{ display: none; animation: fadeIn 0.3s ease; }}
    @keyframes fadeIn {{ from {{ opacity: 0; transform: translateY(10px);} to {{ opacity: 1; transform: translateY(0);} }} }}
    .tab-content.active {{ display: block; }}
    .card {{ background: white; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); transition: box-shadow 0.3s ease; }}
    .card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,0.12); }}
    .card h2 {{ margin: 0 0 20px 0; color: #2C5F2D; font-size: 20px; display: flex; align-items: center; gap: 8px; }}
    .card h2 i {{ font-size: 18px; opacity: 0.7; }}
    .chart-container {{ position: relative; height: 300px; width: 100%; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
    .grid-3 {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 16px; margin-bottom: 20px; }}
    .metric-card {{ background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%); border-left: 4px solid #2C5F2D; padding: 20px; border-radius: 8px; transition: all 0.3s ease; }}
    .metric-card:hover {{ transform: translateX(4px); box-shadow: 0 4px 12px rgba(44, 95, 45, 0.15); }}
    .metric-card strong {{ display: block; color: #1C3F60; margin-bottom: 8px; font-size: 14px; }}
    .metric-card .value {{ font-size: 28px; font-weight: 700; color: #2C5F2D; }}
    .metric-card .subtext {{ font-size: 12px; color: #666; margin-top: 4px; }}
    .data-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    .data-table th, .data-table td {{ border: 1px solid #e5e5e5; padding: 12px; text-align: left; }}
    .data-table thead th {{ background: #2C5F2D; color: white; font-weight: 600; position: sticky; top: 0; z-index: 10; }}
    .data-table tbody tr:nth-child(even) {{ background: #f9f9f9; }}
    .data-table tbody tr:hover {{ background: #e8f5e9; cursor: pointer; }}
    .opportunity-badge {{ display: inline-block; padding: 6px 12px; border-radius: 16px; font-size: 12px; font-weight: 600; }}
    .priority-1 {{ background: #DC2626; color: white; }}
    .priority-2 {{ background: #FF8C00; color: white; }}
    .priority-3 {{ background: #FFD700; color: #333; }}
    @media(max-width: 768px) {{ .kpi-row {{ grid-template-columns: 1fr; }} .grid-2, .grid-3 {{ grid-template-columns: 1fr; }} .tab-navigation {{ overflow-x: auto; }} .dashboard-header {{ padding: 16px; }} .property-title {{ font-size: 20px; }} }}
    .footer {{ text-align: center; padding: 24px; color: #64748b; font-size: 13px; border-top: 1px solid #e2e8f0; margin-top: 40px; }}
    .footer strong {{ color: #2C5F2D; }}
  </style>
</head>
<body>
  <header class="dashboard-header">
    <div class="property-title">{profile['name']}</div>
    <div class="analysis-date">WasteWise Complete Analysis | {(' - '.join(expense_labels)) or 'Period not provided'}</div>

    <div class="kpi-row">
      <div class="kpi-card primary"><div class="kpi-label">Avg Tons/Haul</div><div class="kpi-value">{m['avg_tons_per_haul']:.1f}</div></div>
      <div class="kpi-card primary"><div class="kpi-label">Cost Per Door</div><div class="kpi-value">{fmt_currency(m['avg_cost_per_door'])}</div></div>
      <div class="kpi-card"><div class="kpi-label">Total Spend (period)</div><div class="kpi-value">{fmt_currency(m['total_spend'])}</div></div>
      <div class="kpi-card"><div class="kpi-label">Projected Annual</div><div class="kpi-value">{fmt_currency(m['avg_monthly_spend']*12)}</div></div>
      <div class="kpi-card"><div class="kpi-label">Compactor Util %</div><div class="kpi-value">{(m['avg_tons_per_haul']/8*100 if m['avg_tons_per_haul'] else 0):.0f}%</div></div>
      <div class="kpi-card"><div class="kpi-label">Cost Per Ton</div><div class="kpi-value">{fmt_currency((m['avg_monthly_spend']/ (m['avg_tons_per_haul']*m['hauls_per_month'] or 1)) if m['hauls_per_month'] else 0)}</div></div>
      <div class="kpi-card primary"><div class="kpi-label">Yards Per Door</div><div class="kpi-value">{m['yards_per_door']:.2f}</div></div>
    </div>

    <nav class="tab-navigation">
      <button class="tab-btn active" onclick="openTab('dashboard', this)"><i class="fas fa-chart-line"></i> Dashboard</button>
      <button class="tab-btn" onclick="openTab('expenses', this)"><i class="fas fa-file-invoice-dollar"></i> Expense Analysis</button>
      <button class="tab-btn" onclick="openTab('haul', this)"><i class="fas fa-dolly"></i> Haul Log</button>
      <button class="tab-btn" onclick="openTab('optimization', this)"><i class="fas fa-lightbulb"></i> Optimization</button>
      <button class="tab-btn" onclick="openTab('contract', this)"><i class="fas fa-file-contract"></i> Contract Terms</button>
      <button class="tab-btn" onclick="openTab('regulatory', this)"><i class="fas fa-gavel"></i> Regulatory</button>
    </nav>
  </header>

  <div class="container">
    <div id="dashboard" class="tab-content active">
      <div class="grid-3">
        <div class="metric-card"><strong>Total Compactor Hauls</strong><div class="value">{len(haul)}</div><div class="subtext">{m['hauls_per_month']:.1f} hauls per month (est)</div></div>
        <div class="metric-card"><strong>Avg Days Between Service</strong><div class="value">{profile.get('max_days_between_pickups','-')} days</div><div class="subtext">Threshold ≤14 days</div></div>
        <div class="metric-card"><strong>Potential Annual Savings</strong><div class="value">{fmt_currency(analysis['total_annual_savings'])}</div><div class="subtext">{len(recs)} opportunities prioritized</div></div>
      </div>

      <div class="grid-2">
        <div class="card">
          <h2><i class="fas fa-chart-line"></i> Monthly Expense Trend</h2>
          <div class="chart-container"><canvas id="expenseTrendChart"></canvas></div>
        </div>
        <div class="card">
          <h2><i class="fas fa-weight"></i> Monthly Tonnage Trend</h2>
          <div class="chart-container"><canvas id="tonnageTrendChart"></canvas></div>
        </div>
      </div>

      <div class="card">
        <h2><i class="fas fa-lightbulb"></i> Top Optimization Opportunities</h2>
        <table class="data-table">
          <thead><tr><th>Priority</th><th>Opportunity</th><th>Annual Savings</th><th>ROI</th><th>Complexity</th><th>Timeline</th></tr></thead>
          <tbody>
"""
        for rec in recs:
            html += f"""            <tr>
              <td><span class="opportunity-badge priority-{min(3, rec.priority or 3)}">{rec.priority}</span></td>
              <td><strong>{rec.title}</strong><br><small>{rec.detail}</small></td>
              <td><strong>{fmt_currency(rec.annual_savings)}</strong></td>
              <td>High</td>
              <td>{"Low" if rec.title != "Add Compactor Monitors" else "Medium"}</td>
              <td>{"15-45 days"}</td>
            </tr>
"""
        html += """          </tbody>
        </table>
      </div>
    </div>

    <div id="expenses" class="tab-content">
      <div class="card">
        <h2><i class="fas fa-file-invoice-dollar"></i> Expense Analysis</h2>
        <div style="overflow-x: auto; max-height: 600px;">
          <table class="data-table" id="expenseTable">
            <thead><tr><th>Month</th><th>Invoice #</th><th>Pickup</th><th>Disposal</th><th>Rental</th><th>Contam</th><th>Bulk</th><th>Overages</th><th>Total</th><th>Cost/Door</th><th>Notes</th></tr></thead>
            <tbody>
"""
        units = profile["units"]
        for inv in invoices:
            total = inv.pickup_fees + inv.disposal + inv.rental + inv.contamination + inv.bulk + inv.overages + inv.other
            cost_per_door = total / units if units else 0
            html += f"""              <tr><td>{inv.month}</td><td>{inv.invoice_number}</td><td>{fmt_currency(inv.pickup_fees)}</td><td>{fmt_currency(inv.disposal)}</td><td>{fmt_currency(inv.rental)}</td><td>{fmt_currency(inv.contamination)}</td><td>{fmt_currency(inv.bulk)}</td><td>{fmt_currency(inv.overages)}</td><td>{fmt_currency(total)}</td><td>{fmt_currency(cost_per_door)}</td><td></td></tr>
"""
        html += """            </tbody>
          </table>
        </div>
      </div>
    </div>

    <div id="haul" class="tab-content">
      <div class="card">
        <h2><i class="fas fa-dolly"></i> Haul Log</h2>
        <div style="overflow-x: auto;">
          <table class="data-table">
            <thead><tr><th>Date</th><th>Tons</th><th>Avg Tons/Haul</th><th>Max Days Between Pickups</th><th>Yards/Door</th><th>Notes</th></tr></thead>
            <tbody>
"""
        for h in haul:
            html += f"""              <tr><td>{h.get('pickup_date','')}</td><td>{h.get('tons','')}</td><td>{h.get('avg_tons_per_haul','')}</td><td>{h.get('max_days_between_pickups','')}</td><td>{h.get('yards_per_door','')}</td><td>{h.get('notes','')}</td></tr>
"""
        html += """            </tbody>
          </table>
        </div>
      </div>
    </div>

    <div id="optimization" class="tab-content">
      <div class="card">
        <h2><i class="fas fa-lightbulb"></i> Implementation Roadmap</h2>
        <div style="background: #D1FAE5; border-left: 4px solid #22C55E; padding: 20px; border-radius: 8px; margin-bottom: 24px;">
          <h3 style="color: #059669; margin-bottom: 12px;">Total Savings Potential: {fmt_currency(analysis['total_annual_savings'])} annually</h3>
          <p style="color: #047857; margin: 0;">Top opportunities reduce pickup frequency, contamination, and bulk costs without impacting service quality.</p>
        </div>
        <h3 style="margin: 24px 0 16px; color: #2C5F2D;">Contacts for Implementation</h3>
        <p><strong>Compactor Monitors:</strong> {self.contacts['compactor_monitors']}</p>
        <p><strong>Bulk / Contamination:</strong> {self.contacts['bulk_trash']}</p>
        <p><strong>Primary Contact:</strong> {self.contacts['primary']}</p>
      </div>
    </div>

    <div id="contract" class="tab-content">
      <div class="card">
        <h2><i class="fas fa-file-contract"></i> Contract Terms (Sample)</h2>
        <table class="data-table">
          <thead><tr><th>Clause</th><th>Details</th><th>Risk</th></tr></thead>
          <tbody>
"""
        contract: ContractInfo = analysis["contract"]
        if contract.provided:
            html += f"""            <tr><td>Term & Renewal</td><td>Expires {contract.expiration_date}; {contract.auto_renewal}</td><td>MED</td></tr>
            <tr><td>Rate Increases</td><td>{contract.rate_increase_terms}</td><td>LOW</td></tr>
            <tr><td>Termination</td><td>{contract.termination_notice_days}-day notice; {contract.termination_clause}</td><td>MED</td></tr>
            <tr><td>Liability</td><td>{contract.liability_text}</td><td>MED</td></tr>
            <tr><td>Service Level</td><td>{contract.service_level_text}</td><td>LOW</td></tr>
            <tr><td>Force Majeure</td><td>{contract.force_majeure_text}</td><td>LOW</td></tr>
            <tr><td>Indemnification</td><td>{contract.indemnification_text}</td><td>MED</td></tr>
"""
        else:
            html += "<tr><td>Contract</td><td>No contract provided</td><td>HIGH</td></tr>"
        html += """          </tbody>
        </table>
      </div>
    </div>

    <div id="regulatory" class="tab-content">
      <div class="card">
        <h2><i class="fas fa-gavel"></i> Regulatory Compliance ({profile.get('location','N/A')})</h2>
        <div style="background:#fef9c3; border-left:4px solid #eab308; padding:16px; border-radius:8px; margin-bottom:16px;">
          <strong>Confidence: LOW</strong> — placeholder. Update with official sources and penalties; list 3-5 licensed haulers.
        </div>
        <ul style="margin-left: 20px; line-height: 1.7;">
          <li>Recycling: Confirm universal recycling requirements, frequency, and signage.</li>
          <li>Organics: Check any composting mandates and effective dates.</li>
          <li>Penalties: Document fine amounts and enforcement agency.</li>
          <li>Licensed haulers: Provide 3–5 with contact info.</li>
        </ul>
      </div>
    </div>
  </div>

  <div class="footer">
    <p><strong>WasteWise Complete Analysis</strong> by <strong>THE Trash Hub</strong></p>
    <p>Report generated for {profile['name']} | {profile['units']} units | {profile['property_type'].title()} Property</p>
  </div>

<script>
function openTab(tabId, btnElement) {{
  document.querySelectorAll('.tab-content').forEach(tab => tab.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
  document.getElementById(tabId).classList.add('active');
  btnElement.classList.add('active');
}}

const expenseLabels = {json.dumps(expense_labels)};
const expenseTotals = {json.dumps(expense_totals)};
const tonnageLabels = {json.dumps(tonnage_labels)};
const tonnageTotals = {json.dumps(tonnage_totals)};

const expenseCtx = document.getElementById('expenseTrendChart').getContext('2d');
new Chart(expenseCtx, {{
  type: 'line',
  data: {{ labels: expenseLabels, datasets: [{{ label: 'Total Monthly Cost', data: expenseTotals, borderColor: '#2C5F2D', backgroundColor: 'rgba(44,95,45,0.1)', fill: true, tension: 0.4, borderWidth: 3, pointRadius: 5, pointHoverRadius: 7, pointBackgroundColor: '#2C5F2D', pointBorderColor: '#fff', pointBorderWidth: 2 }}] }},
  options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }}, tooltip: {{ backgroundColor: 'rgba(0,0,0,0.8)', padding: 12, callbacks: {{ label: (ctx) => 'Total: $' + ctx.parsed.y.toLocaleString() }} }} }}, scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: val => '$' + val.toLocaleString() }} }} }} }}
}});

const tonnageCtx = document.getElementById('tonnageTrendChart').getContext('2d');
new Chart(tonnageCtx, {{
  type: 'bar',
  data: {{ labels: tonnageLabels, datasets: [{{ label: 'Total Tons', data: tonnageTotals, backgroundColor: 'rgba(74,124,78,0.8)', borderColor: '#4A7C4E', borderWidth: 2, borderRadius: 6 }}] }},
  options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }}, tooltip: {{ backgroundColor: 'rgba(0,0,0,0.8)', padding: 12, callbacks: {{ label: (ctx) => 'Tonnage: ' + (ctx.parsed.y ?? 0).toFixed(1) + ' tons' }} }} }}, scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: val => val.toFixed(1) + ' tons' }} }} }} }}
}});
</script>
</body>
</html>
"""
        path.write_text(html, encoding="utf-8")
        return path


def load_payload(path: Path) -> Dict:
    with path.open() as f:
        return json.load(f)


def run_end_to_end(input_path: Path, excel_path: Path, html_path: Path) -> Dict:
    agent = WasteWiseAgent()
    payload = load_payload(input_path)
    analysis = agent.analyze(payload)
    agent.build_excel_workbook(analysis, excel_path)
    agent.build_html_dashboard(analysis, html_path)
    return {
        "excel": str(excel_path),
        "html": str(html_path),
        "annual_savings": analysis["total_annual_savings"],
    }


if __name__ == "__main__":
    base = Path(__file__).parent
    input_path = base / "sample_data.json"
    excel_path = base / "wastewise_output.xlsx"
    html_path = base / "wastewise_dashboard.html"
    result = run_end_to_end(input_path, excel_path, html_path)
    print(json.dumps(result, indent=2, ensure_ascii=True))
